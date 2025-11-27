import os
import yaml
import glob
import time
import pandas as pd
from loguru import logger
from typing import Union, List, Any
from unidecode import unidecode
from openpyxl import load_workbook


def Registro_tiempo(original_func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = original_func(*args, **kwargs)
        end_time = time.time()
        execution_time = end_time - start_time
        logger.info(
            f"Tiempo de ejecución de {original_func.__name__}: {execution_time} segundos"
        )
        return result

    return wrapper


def Procesar_configuracion(nom_archivo_configuracion: str) -> dict:
    """Lee un archivo YAML de configuración para un proyecto.

    Args:
        nom_archivo_configuracion (str): Nombre del archivo YAML que contiene
            la configuración del proyecto.

    Returns:
        dict: Un diccionario con la información de configuración leída del archivo YAML.
    """
    try:
        with open(nom_archivo_configuracion, "r", encoding="utf-8") as archivo:
            configuracion_yaml = yaml.safe_load(archivo)
        logger.success("Proceso de obtención de configuración satisfactorio")
    except Exception as e:
        logger.critical(f"Proceso de lectura de configuración fallido {e}")
        raise e

    return configuracion_yaml


@Registro_tiempo
def Lectura_insumos_excel(
    path: str, nom_insumo: str, nom_Hoja: str, cols: Union[int, list]
) -> pd.DataFrame:
    """Lee archivos de Excel con cualquier extensión y carga los datos de una hoja específica.

    Lee el archivo especificado por `nom_insumo` ubicado en la ruta `path` y carga los datos de la hoja
    especificada por `nom_Hoja`. Selecciona solo las columnas indicadas por `cols`.

    Args:
        path (str): Ruta de la carpeta donde se encuentra el archivo.
        nom_insumo (str): Nombre del archivo con extensión.
        nom_Hoja (str): Nombre de la hoja del archivo que se quiere leer.
        cols (int): Número de columnas que se desean cargar.

    Returns:
        pd.DataFrame: Dataframe que contiene los datos leídos del archivo Excel.

    Raises:
        Exception: Si ocurre un error durante el proceso de lectura del archivo.
    """
    base_leida = None

    try:
        logger.info(f"Inicio lectura {nom_insumo} Hoja {nom_Hoja}")
        base_leida = pd.read_excel(
            path + nom_insumo,
            sheet_name=nom_Hoja,
            usecols=list(range(0, cols)),
            dtype=str,
        )

        logger.success(
            f"Lectura de {nom_insumo} Hoja: {nom_Hoja} realizada con éxito"
        )  # Se registrará correctamente con el método "success"
    except Exception as e:
        logger.error(f"Proceso de lectura fallido: {e}")
        raise Exception

    return base_leida


@Registro_tiempo
def leer_carpeta_de_archivos_excel(carpeta):
    """
    Lee todos los archivos de Excel en una carpeta como `str`.

    Args:
        carpeta: Carpeta que contiene los archivos de Excel.

    Returns:
        Una lista de DataFrames con los datos de todos los archivos de Excel.
    """

    logger.info(f"Obtenemos los archivos de la carpeta: {carpeta}")
    archivos = glob.glob(os.path.join(carpeta, "*.xlsx"))
    df_list = []
    try:
        logger.info(f"Inicio lectura archivos")
        for archivo in archivos:
            df = pd.read_excel(archivo, dtype=str)
            logger.info("Se está leyendo el archivo {}".format(archivo))
            df_list.append(df)

    except Exception as e:
        logger.critical(e)
        raise e

    return df_list


@Registro_tiempo
def pd_left_merge(
    base_left: pd.DataFrame, base_right: pd.DataFrame, key: str
) -> pd.DataFrame:
    """Función que retorna el left join de dos dataframe de pandas.

    Args:
        base_left (pd.DataFrame): Dataframe que será la base del join.
        base_right (pd.DataFrame): Dataframe del cuál se extraerá la información complementaria.
        key (str): Llave mediante la cual se va a realizar el merge o join.

    Returns:
        pd.DataFrame: Dataframe con el merge de las dos fuentes de datos.
    """

    # Validar que base_left y base_right sean DataFrames de pandas
    if not isinstance(base_left, pd.DataFrame):
        raise ValueError("El argumento base_left no es un DataFrame de pandas")
    if not isinstance(base_right, pd.DataFrame):
        raise ValueError("El argumento base_right no es un DataFrame de pandas")

    base = None

    try:
        base = pd.merge(left=base_left, right=base_right, how="left", on=key)
        logger.success("Proceso de merge satisfactorio")
    except pd.errors.MergeError as e:
        logger.critical(f"Proceso de merge fallido: {e}")
        raise e

    return base


@Registro_tiempo
def Cargar_insumo(ruta: str, base_config: str, num_hojas_consulta: int) -> list:
    """
    Carga consultas a partir de una configuración base en un directorio específico.

    Args:
        ruta (str): Ruta al directorio donde se encuentran los archivos de consulta.
        base_config (dict): Configuración base para las consultas, incluyendo nombre de archivo,
                           hojas y columnas.
        num_hojas_consulta (int): Número de hojas de consulta a cargar.

    Returns:
        list: Lista de consultas cargadas.

    Example:
        base_no_ds = {
            "file_name": "archivo_no_ds.xlsx",
            "sheet": ["hoja1", "hoja2"],
            "cols": [4, 5] => Indica que "hoja1" tiene 4 columnas, la "hoja2" 5
        }
        num_hojas_no_ds = len(base_no_ds["sheet"])
        consultas_no_ds = cargar_consultas(ruta, base_no_ds, num_hojas_no_ds)

    """
    try:
        consultas = []
        for i in range(0, num_hojas_consulta):
            consulta = Lectura_insumos_excel(
                path=ruta,
                nom_insumo=base_config["file_name"],
                nom_Hoja=base_config["sheet"][i],
                cols=base_config["cols"][i],
            )
            consultas.append(consulta)
    except Exception as e:
        logger.error(f"Proceso de lectura fallido: {e}")
        raise e

    return consultas


def Crear_diccionario_con_listas(dataframe: pd.DataFrame, col_clave: str) -> dict:
    """
    Crea un diccionario donde las claves son los valores de la columna 1
    y los valores son listas de los valores de las columnas 3 en adelante.

    Args:
        dataframe (pd.DataFrame): El DataFrame de entrada con las columnas deseadas.

    Returns:
        dict: Un diccionario con listas de valores de las columnas 1 en adelante.
        OJO: EL orden de las columnas importa. Ya que la columa 0 se usa como clave.
        Y no entra en el dict.
    """
    diccionario = {}
    try:
        for index, row in dataframe.iterrows():
            clave = row[col_clave]
            valores = list(row.iloc[1:])

            diccionario[clave] = valores

    except KeyError as e:
        logger.critical(f"Creación del diccionario fallida debido a una KeyError: {e}")
    except Exception as e:
        logger.critical(f"Creación del diccionario fallida. Error: {e}")

    return diccionario


import loguru

logger = loguru.logger


def extraer_sublista_hasta_elemento(lista: list, elemento_final: str) -> list:
    """
    Extrae una sublista desde el principio de la lista original hasta el índice del elemento especificado.

    Args:
      lista: Lista original de la que se extraerá la sublista.
      elemento_final: Elemento que marca el final de la sublista.

    Returns:
      Sublista que contiene los elementos desde el principio de la lista original hasta el elemento especificado.
    """

    try:
        indice_final = lista.index(elemento_final)
        sublista = lista[: indice_final + 1]
        logger.success("Sublista extraída correctamente.")
        return sublista
    except ValueError as e:
        logger.critical(f"Error al extraer la sublista: {e}")
        raise e


def encontrar_llave_por_valor(diccionario, valor):
    """
    Encuentra la llave correspondiente a un valor dado en un diccionario.

    Parámetros:
    - diccionario (dict): El diccionario en el que buscar la llave.
    - valor: El valor para el cual encontrar la llave.

    Retorna:
    - str: La llave correspondiente al valor.

    Ejemplo:
    >>> mi_diccionario = {'a': 1, 'b': 2, 'c': 3}
    >>> mi_valor = 2
    >>> encontrar_llave_por_valor(mi_diccionario, mi_valor)
    'b'
    """
    try:
        llave_correspondiente = next(
            llave for llave, v in diccionario.items() if v == valor
        )
        return llave_correspondiente
    except StopIteration:
        logger.critical(f"No se encontró la llave para el valor {valor}")
        raise ValueError(
            f"No se encontró la llave para el valor {valor} en el diccionario."
        )


@Registro_tiempo
def Eliminar_acentos(df: pd.DataFrame) -> pd.DataFrame:
    """
    Elimina las tildes y caracteres acentuados de todas las columnas de un DataFrame.

    Args:
        df (pd.DataFrame): El DataFrame en el que se eliminarán los acentos.

    Returns:
        pd.DataFrame: El DataFrame modificado con los acentos eliminados.
    """
    try:
        logger.info("Iniciando proceso de eliminación de acentos...")
        for columna in df.columns:
            # Verifica si la columna es de tipo objeto (texto)
            if df[columna].dtype == "object":
                # Aplica unidecode a cada valor de la columna y asigna los resultados de nuevo a la columna
                df[columna] = df[columna].apply(
                    lambda x: unidecode(x) if pd.notna(x) else x
                )
        logger.success("Proceso de eliminación de acentos completado con éxito.")
        return df
    except Exception as e:
        # Manejo de excepciones: registra un mensaje crítico en lugar de imprimir el error
        logger.critical(f"Error en la función eliminar_acentos: {str(e)}")
        return df


@Registro_tiempo
def Exportar_dataframes_mismo_excel(
    dataframes: dict, nombre_archivo: str
) -> pd.ExcelWriter:
    """Función que toma las claves y los valores de un diccioanrio donde, y crea un objeto ExcelWriter que a su vez escribe un archivo de excel con la información de
    dataframes:

    Args:
        dataframes (dict): Diccionario donde:
            claves: Nombre de la hoja asignada para cada dataframe.
            valores: Dataframes correspondientes.
        nombre_archivo: Nombre que tendrá el archivo final.

    Returns:
        pd.ExcelWriter: objeto para escribir el resultado final, archivo.xlsx
        bajo el nombre de "nombre_archivo"
    """
    try:
        logger.info(f"Inico creacion *.xlsx => {nombre_archivo}")
        with pd.ExcelWriter(nombre_archivo) as writer:
            for key, value in dataframes.items():
                logger.info(f"Procesando: {key}")
                value.to_excel(writer, sheet_name=key, index=False)
                logger.success(f"{key} añadido correctamente")
    except Exception as e:
        logger.critical(f"Procedimiento fallido con: {key}")
        raise e

from openpyxl import load_workbook

def Leer_nombres_columnas_excel(archivo: str, hoja: str = None) -> list:
    """
    Lee los nombres de las columnas de la primera fila de un archivo Excel
    sin cargar todo el archivo en la memoria utilizando la biblioteca openpyxl.

    Args:
        archivo (str): Ruta al archivo Excel del cual leer los nombres de las columnas.
        hoja (str, opcional): Nombre de la hoja de Excel de la que se leerán las columnas.
                              Si se omite, se seleccionará la hoja activa por defecto.

    Returns:
        Una lista de strings con los nombres de las columnas del archivo Excel.

    Ejemplo:
        nombres_columnas = leer_nombres_columnas_excel('datos.xlsx', 'Hoja1')
        print(nombres_columnas)
    """
    # Cargar el libro de trabajo en modo solo lectura
    wb = load_workbook(filename=archivo, read_only=True)
    
    # Seleccionar la hoja por nombre o por defecto la hoja activa
    if hoja:
        sheet = wb[hoja]
    else:
        sheet = wb.active

    # Obtener los nombres de las columnas de la primera fila
    columnas = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Cerrar el libro de trabajo
    wb.close()

    return columnas


def exportar_a_excel(
    ruta_guardado: str, df: pd.DataFrame, nom_hoja: str, index: bool = False
) -> None:
    """
    Exporta un dataframe de pandas a un archivo excel en la ruta especificada.

    Args:
        ruta_guardado: Ruta donde se guardará el archivo excel.
        df: Dataframe de pandas que se exportará.
        nom_hoja: Nombre de la hoja de cálculo donde se exportará el dataframe.
        index: Indica si se debe incluir el índice del dataframe en el archivo excel.

    Returns:
        None.

    Raises:
        FileNotFoundError: Si la ruta de guardado no existe.
    """

    # Comprobar que la ruta de guardado existe
    try:
        logger.info(f"Exportando a excel: {nom_hoja}")
        df.to_excel(
            ruta_guardado + nom_hoja + ".xlsx", sheet_name=nom_hoja, index=index
        )
    except Exception as e:
        raise Exception


from loguru import logger


def verificar_existencia_columnas(df, columnas: Union[list, set]):
    """
    Verifica si todas las columnas especificadas existen en el DataFrame.

    Args:
        df: DataFrame a verificar.
        columnas: Lista o conjunto de nombres de columnas a verificar.

    Raises:
        KeyError: Si alguna columna no existe en el DataFrame.
        ValueError: Si df no es un DataFrame.
    """
    try:
        # Verificar si df es un DataFrame
        if not isinstance(df, pd.DataFrame):
            raise ValueError("El argumento 'df' no es un DataFrame.")

        # Verificar la existencia de columnas
        columnas_no_existentes = set(columnas) - set(df.columns)
        if columnas_no_existentes:
            raise KeyError(
                f"Las siguientes columnas no existen en el DataFrame: {', '.join(columnas_no_existentes)}"
            )
            
        # Si no se lanzó ninguna excepción, retornar True
        return True
    
    except ValueError as ve:
        logger.critical(f"Error: {ve}")
        raise ve
    except KeyError as ke:
        logger.critical(f"Error al verificar las columnas: {ke}")
        raise ke
    except Exception as e:
        logger.critical(f"Error inesperado: {e}")
        raise e
